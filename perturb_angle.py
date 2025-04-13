"""
Angle-Based Perturbation Analysis Module.

This module implements angle-based perturbation analysis for understanding
the behavior of our Cerebral Palsy prediction model.
"""

import os
import math
import json
import pickle
import copy
import random
from collections import defaultdict
from typing import List, Dict, Tuple, Any, Optional, Union

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d
from openpyxl import load_workbook
from openpyxl.styles import Font

# Import prediction function from private module
from predict import predict, predict_with_pvb
from utils.predict_helpers import read_csv_to_array


# Constants
BODY_PARTS = [
    'head_top', 'nose', 'right_ear', 'left_ear', 'upper_neck',
    'right_shoulder', 'right_elbow', 'right_wrist', 'thorax',
    'left_shoulder', 'left_elbow', 'left_wrist', 'pelvis',
    'right_hip', 'right_knee', 'right_ankle', 'left_hip',
    'left_knee', 'left_ankle'
]

CP_RISK_THRESHOLD = 0.35025954  # Risk threshold for CP classification
CAM_THRESHOLD = 0.2915886665789474  # CAM threshold

# Joint groupings
TOP_K_LOW_RISK = [7, 11, 14, 15]  # Identified important joints
NON_TOP_K_LOW_RISK = [0, 1, 2, 3, 4, 5, 6, 9, 10, 12, 13, 16, 17, 18]  # Other joints

# Raw body part names with x,y coordinates
BODY_PARTS_RAW = [
    'head_top_x', 'head_top_y', 'nose_x', 'nose_y', 'right_ear_x', 'right_ear_y',
    'left_ear_x', 'left_ear_y', 'upper_neck_x', 'upper_neck_y', 'right_shoulder_x',
    'right_shoulder_y', 'right_elbow_x', 'right_elbow_y', 'right_wrist_x', 'right_wrist_y',
    'thorax_x', 'thorax_y', 'left_shoulder_x', 'left_shoulder_y', 'left_elbow_x',
    'left_elbow_y', 'left_wrist_x', 'left_wrist_y', 'pelvis_x', 'pelvis_y',
    'right_hip_x', 'right_hip_y', 'right_knee_x', 'right_knee_y', 'right_ankle_x',
    'right_ankle_y', 'left_hip_x', 'left_hip_y', 'left_knee_x', 'left_knee_y',
    'left_ankle_x', 'left_ankle_y'
]

# Mapping from raw indices to joint indices
BODY_PARTS_RAW_NUM = [
    0, 0, 1, 1, 2, 2, 3, 3, 4, 4, 5, 5, 6, 6, 7, 7, 8, 8, 9, 9, 10, 10, 11, 11,
    12, 12, 13, 13, 14, 14, 15, 15, 16, 16, 17, 17, 18, 18
]

# Parent-child joint relationships for angle calculations
ADJACENCY_MAP = {
    0: 1, 2: 1, 3: 1, 1: 4, 9: 8, 10: 9, 11: 10, 5: 8, 6: 5, 7: 6,
    4: 8, 12: 8, 16: 12, 17: 16, 18: 17, 13: 12, 14: 13, 15: 14
}


def load_video_fps_dict(json_path: str) -> Dict[str, float]:
    """
    Load video frame rate dictionary from JSON file and modify keys.
    
    Args:
        json_path: Path to the JSON file containing video FPS information
        
    Returns:
        Dictionary mapping CSV filenames to video frame rates
    """
    with open(json_path, 'r') as json_file:
        video_fps_dict = json.load(json_file)
    
    # Add prefix 'tracked_' and suffix '.csv' to all keys
    return {'tracked_' + key + '.csv': value for key, value in video_fps_dict.items()}


def get_frame_range_for_window(
    window_number: int,
    frame_rate: float,
    num_frames: int,
    pred_interval_seconds: float = 5
) -> Tuple[Optional[int], Optional[int]]:
    """
    Calculate the start and end frame indices for a specific window.
    
    Args:
        window_number: Index of the window to analyze
        frame_rate: Frame rate of the video
        num_frames: Total number of frames in the video
        pred_interval_seconds: Size of prediction window in seconds
        
    Returns:
        Tuple with start and end frame indices, or (None, None) if invalid
    """
    window_size = int(frame_rate * pred_interval_seconds)
    step_size = int(window_size / 2)
    
    # Compute start and end frame
    start_frame = window_number * step_size
    end_frame = start_frame + window_size - 1
    
    # Ensure that the start frame doesn't go beyond the total number of frames
    if start_frame >= num_frames:
        return None, None  # Invalid window number
    
    # Ensure that the end frame does not exceed the total number of frames
    if end_frame >= num_frames:
        end_frame = num_frames - 1
    
    # Adjust the window if it's the last valid window
    # Check if this window is the last full window or partial window
    total_windows = math.floor((num_frames - step_size) / step_size) + 1
    if window_number == total_windows - 1:
        start_frame = num_frames - window_size if num_frames >= window_size else 0
        end_frame = num_frames - 1
    
    return start_frame, end_frame


def decision_logic(data: np.ndarray, threshold: float) -> int:
    """
    Apply decision logic to classify results based on risk threshold.
    
    Args:
        data: Array of predictions
        threshold: Risk threshold for classification
        
    Returns:
        Uncertainty level (0: low risk, 1: uncertain-low, 2: uncertain-high, 3: high risk)
    """
    median = np.median(data, axis=0)
    lower_bound, upper_bound = np.percentile(data, [25, 75], axis=0)
    
    if upper_bound < threshold:
        uncertainty = 0  # Low risk
    elif lower_bound > threshold:
        uncertainty = 3  # High risk
    else:
        if median >= threshold:
            uncertainty = 2  # Uncertain but more likely high risk
        else:
            uncertainty = 1  # Uncertain but more likely low risk
            
    return uncertainty


def calculate_angle_changes(pose_data: np.ndarray) -> np.ndarray:
    """
    Calculate frame-to-frame angle changes between connected body parts.
    Frame 0 will have 0 changes (no previous frame).
    Frame 1 will have changes from frame 0 to 1.
    Frame i will have changes from frame i-1 to i.
    
    Args:
        pose_data: Numpy array of shape (n, 38) where n is number of frames
                  and 38 corresponds to x,y coordinates of 19 body points
    
    Returns:
        Numpy array of shape (n, 19) containing angle changes for each joint
    """
    pose_data = pose_data.astype(float)
    n_frames = pose_data.shape[0]
    
    # Initialize output array with zeros
    angle_changes = np.zeros((n_frames, 19))
    
    def calculate_angle(point1: Tuple[float, float], point2: Tuple[float, float]) -> float:
        """Calculate angle between two points relative to horizontal."""
        dx = point2[0] - point1[0]
        dy = point2[1] - point1[1]
        return np.degrees(np.arctan2(dy, dx))
    
    # Process each frame starting from frame 1
    # frame i will store changes from frame i-1 to i
    for frame in range(1, n_frames):
        # Process each child-parent relationship
        for child, parent in ADJACENCY_MAP.items():
            # Get coordinates for previous frame (i-1)
            child_x_prev = pose_data[frame-1, child * 2]
            child_y_prev = pose_data[frame-1, child * 2 + 1]
            parent_x_prev = pose_data[frame-1, parent * 2]
            parent_y_prev = pose_data[frame-1, parent * 2 + 1]
            
            # Get coordinates for current frame (i)
            child_x_curr = pose_data[frame, child * 2]
            child_y_curr = pose_data[frame, child * 2 + 1]
            parent_x_curr = pose_data[frame, parent * 2]
            parent_y_curr = pose_data[frame, parent * 2 + 1]
            
            # Calculate angles in previous and current frame
            angle_prev = calculate_angle(
                (parent_x_prev, parent_y_prev),
                (child_x_prev, child_y_prev)
            )
            angle_curr = calculate_angle(
                (parent_x_curr, parent_y_curr),
                (child_x_curr, child_y_curr)
            )
            
            # Calculate angle change and store in current frame
            angle_change = angle_curr - angle_prev
            
            # Normalize angle change to be between -180 and 180 degrees
            angle_change = (angle_change + 180) % 360 - 180
            
            # Store the angle change in the output array
            angle_changes[frame, child] = angle_change
    
    # Frame 0 remains 0 (initialized with zeros)
    return angle_changes


def analyze_angle_changes(
    angle_changes_list: List[np.ndarray]
) -> Tuple[pd.DataFrame, List[List[float]]]:
    """
    Analyze angle changes across multiple skeleton files and calculate percentiles, min, and max.
    
    Args:
        angle_changes_list: List of numpy arrays, each of shape (n_i, 19) where n_i
                          is the number of frames for skeleton i
    
    Returns:
        - Pandas DataFrame with joints as rows and statistics as columns
        - List of lists containing the same data
    """
    # Initialize dictionary to store all non-zero angle changes for each joint
    joint_angles = {i: [] for i in range(19)}
    
    # Gather all non-zero angle changes for each joint
    for angles in angle_changes_list:
        for joint in range(19):
            # Get absolute values of non-zero angles for this joint
            joint_data = np.abs(angles[:, joint])
            non_zero_angles = joint_data[joint_data > 0]
            joint_angles[joint].extend(non_zero_angles)
    
    # Calculate percentiles for each joint
    percentiles = [5, 25, 50, 75, 95]
    results = []
    
    for joint in range(19):
        if joint_angles[joint]:  # Check if we have any non-zero angles
            stats = np.percentile(joint_angles[joint], percentiles)
            min_val = np.min(joint_angles[joint])
            max_val = np.max(joint_angles[joint])
            results.append([joint] + [min_val] + list(stats) + [max_val])
        else:
            # For joints with no non-zero angles (like joint 18)
            results.append([joint] + [np.nan] * (len(percentiles) + 2))
    
    # Create DataFrame
    df = pd.DataFrame(
        results,
        columns=['Joint Number', 'Min', 'P5', 'P25', 'P50', 'P75', 'P95', 'Max']
    )
    
    # Format the numbers to 6 decimal places
    for col in ['Min', 'P5', 'P25', 'P50', 'P75', 'P95', 'Max']:
        df[col] = df[col].apply(lambda x: f'{x:.6f}' if pd.notnull(x) else 'N/A')
    
    # Create list version (without formatting)
    list_results = []
    for joint in range(19):
        if joint_angles[joint]:
            stats = np.percentile(joint_angles[joint], percentiles)
            min_val = np.min(joint_angles[joint])
            max_val = np.max(joint_angles[joint])
            list_results.append([joint] + [min_val] + list(stats) + [max_val])
        else:
            list_results.append([joint] + [np.nan] * (len(percentiles) + 2))
    
    return df, list_results


def perturb_angle(
    arr: np.ndarray,
    mult_factors: np.ndarray,
    j: List[int],
    mode: str,
    debug: bool = False
) -> np.ndarray:
    """
    Perturbs joint angles by modifying the angle between parent and child joints.
    
    The process involves:
    1. Translating parent to origin
    2. Perturbing angle
    3. Translating back to parent's position
    
    Args:
        arr: Pose data array of shape (frames, 38)
        mult_factors: Multiplication factors for angle changes
        j: List of joints to perturb
        mode: 'min' to decrease angles or 'max' to increase
        debug: Whether to print debug information
        
    Returns:
        Perturbed pose data array
    """
    arr = arr.astype(float)
    perturbed_arr = np.copy(arr)
    total_frames = arr.shape[0]

    for frame in range(1, total_frames):
        # Process each joint to perturb
        for joint in j:
            factor = mult_factors[joint]
            parent = ADJACENCY_MAP.get(joint)
            
            if factor != 1 and not np.isnan(factor) and parent is not None:
                # Get coordinate indices
                joint_x = joint * 2
                joint_y = joint * 2 + 1
                parent_x = parent * 2
                parent_y = parent * 2 + 1
    
                # Get original positions for current frame (joint and parent)
                curr_frame_jx = arr[frame, joint_x]
                curr_frame_jy = arr[frame, joint_y]
                curr_frame_px = arr[frame, parent_x]
                curr_frame_py = arr[frame, parent_y]

                # Get original positions for previous frame (joint and parent)
                prev_frame_jx = arr[frame-1, joint_x]
                prev_frame_jy = arr[frame-1, joint_y]
                prev_frame_px = arr[frame-1, parent_x]
                prev_frame_py = arr[frame-1, parent_y]

                # Get angle for current frame (joint to parent)
                curr_frame_dx = curr_frame_jx - curr_frame_px
                curr_frame_dy = curr_frame_jy - curr_frame_py
                curr_frame_angle = np.degrees(np.arctan2(curr_frame_dy, curr_frame_dx))

                # Get angle for previous frame (joint to parent)
                prev_frame_dx = prev_frame_jx - prev_frame_px
                prev_frame_dy = prev_frame_jy - prev_frame_py
                prev_frame_angle = np.degrees(np.arctan2(prev_frame_dy, prev_frame_dx))

                # Get angle delta
                angle_change = curr_frame_angle - prev_frame_angle

                if mode == 'max':
                    if np.abs(angle_change) < np.abs(angle_change*factor):
                        angle_change = angle_change*factor
                        if debug:
                            print('angle change is increased')
                else:  # mode == 'min'
                    if np.abs(angle_change) > np.abs(angle_change*factor):
                        angle_change = angle_change*factor
                        if debug:
                            print('angle change is decreased')

                curr_frame_angle = angle_change + prev_frame_angle
                new_angle = np.radians(curr_frame_angle)
                    
                # 1. translate child wrt new parent while maintaining length
                # Get original length (subtract parent position to child)
                rel_x = curr_frame_jx - curr_frame_px
                rel_y = curr_frame_jy - curr_frame_py
                orig_length = np.sqrt(rel_x**2 + rel_y**2)
  
                # get new/perturbed parent location
                new_parent_x = perturbed_arr[frame, parent_x]
                new_parent_y = perturbed_arr[frame, parent_y]
    
                # translate
                translated_x = orig_length*np.cos(new_angle) + new_parent_x
                translated_y = orig_length*np.sin(new_angle) + new_parent_y
                perturbed_arr[frame, joint_x] = translated_x
                perturbed_arr[frame, joint_y] = translated_y
    
                if debug:
                    # Compute original distance
                    orig_distance = np.sqrt(
                        (arr[frame, joint_x] - arr[frame, parent_x])**2 + 
                        (arr[frame, joint_y] - arr[frame, parent_y])**2
                    )
                    
                    # Compute the new distance
                    new_distance = np.sqrt(
                        (perturbed_arr[frame, joint_x] - perturbed_arr[frame, parent_x])**2 + 
                        (perturbed_arr[frame, joint_y] - perturbed_arr[frame, parent_y])**2
                    )
                        
                    diff = abs(orig_distance - new_distance)
                    if round(diff, 6) > 1e-6:
                        print('joint: ', joint)
                        print('diff: ', diff)
                        print(f"orig child: ({arr[frame, joint_x]},{arr[frame, joint_y]})")
                        print(f'orig parent: ({arr[frame, parent_x]},{arr[frame, parent_y]})')
                        print(f'new child: ({perturbed_arr[frame, joint_x]},{perturbed_arr[frame, joint_y]})')
                        print(f'new parent: ({perturbed_arr[frame, parent_x]},{perturbed_arr[frame, parent_y]})')
                        print('new angle: ', new_angle)

    return perturbed_arr


def update_list(
    perturb_list: List[Dict[str, Any]],
    filename: str,
    window_number: int,
    factor: float,
    new_pred: float,
    step: str
) -> None:
    """
    Update perturbation results in the results list.
    
    Args:
        perturb_list: List of perturbation results
        filename: Video filename
        window_number: Window index
        factor: Speed factor applied
        new_pred: New prediction value
        step: Direction of perturbation ('min' or 'max')
    """
    # Check if the entry already exists
    for entry in perturb_list:
        if entry['filename'] == filename and entry['window'] == window_number:
            # If found, update the existing entry
            key = '*' + step
            factor_str = f'{factor:.2f}' if factor < 1 else f'{int(factor)}'
            entry[f'{factor_str}{key}'] = new_pred
            return


def save_to_excel_with_formatting(
    data: List[Dict[str, Any]],
    filename: str,
    threshold: float
) -> None:
    """
    Save data to Excel with conditional formatting for risk values.
    
    Args:
        data: List of dictionaries with perturbation results
        filename: Output filename without extension
        threshold: Risk threshold for highlighting
    """
    # Convert list of dictionaries to DataFrame
    df = pd.DataFrame(data)
    
    # Save DataFrame to Excel
    excel_file = f"{filename}.xlsx"
    df.to_excel(excel_file, index=False)
    
    # Load the workbook and select the active sheet
    wb = load_workbook(excel_file)
    ws = wb.active

    # Apply conditional formatting: if value < threshold, turn the text red
    red_font = Font(color="FF0000")  # Red text
    for col in ws.iter_cols(min_row=2, min_col=2, max_col=ws.max_column):
        for cell in col:
            if isinstance(cell.value, (int, float)) and cell.value < threshold:
                cell.font = red_font

    # Save the workbook with formatting
    wb.save(excel_file)


def run_angle_perturbation(
    filename_windows: List[Dict[str, Any]],
    joints_to_analyze: List[int],
    is_top_k: bool,
    data_path: str,
    output_filename: str,
    min_max_angles: List[List[float]],
    video_fps_dict: Dict[str, float]
) -> None:
    """
    Run angle-based perturbation analysis for specified joint groups.
    
    Args:
        filename_windows: List of window data (filename and window number)
        joints_to_analyze: List of joint indices to perturb
        is_top_k: Whether analyzing top-k (True) or non-top-k (False) joints
        data_path: Path to tracking data files
        output_filename: Base filename for output files
        min_max_angles: Reference angle statistics from training data
        video_fps_dict: Dictionary with video frame rates
    """
    perturb_results = []
    pred_seconds = 5
    factors_min = [0.2, 0.25, 1/3, 0.5, 1]
    factors_max = [1, 2, 3, 4, 5]
    
    for window_data in filename_windows:
        filename = window_data['filename']
        window_number = int(window_data['window'])
        
        # Load tracking data for this file
        csv_path = os.path.join(data_path, filename)
        tracking_coords = np.array(read_csv_to_array(csv_path))
        num_frames = tracking_coords.shape[0]
        frame_rate = video_fps_dict.get(filename)
        
        if frame_rate is None:
            print(f"Warning: No frame rate found for {filename}, skipping.")
            continue
        
        # Get frame range for this window
        start_frame, end_frame = get_frame_range_for_window(
            window_number, frame_rate, num_frames, pred_seconds
        )
        
        if start_frame is None:
            print(f"Invalid window {window_number} for {filename}, skipping.")
            continue
            
        print(f"Processing window {window_number} from {filename} (frames {start_frame} to {end_frame})")
        
        # Extract frames for this window
        frames_in_window = tracking_coords[start_frame:end_frame + 1]

        # Get original prediction
        _, orig_preds, _ = predict_with_pvb(
            tracking_coords=frames_in_window, 
            body_parts=BODY_PARTS, 
            frame_rate=frame_rate, 
            total_frames=frames_in_window.shape[0],
            pred_frame_rate=30.0,
            pred_interval_seconds=2.5,
            window_stride=2,
            num_models=10,
            num_portions=7,
            prediction_threshold=CP_RISK_THRESHOLD,
            xai_technique='cam'
        )
        orig_pred = np.median(orig_preds)

        # Calculate angle changes for this window
        angle_changes = calculate_angle_changes(frames_in_window)
        angle_changes_list = [angle_changes]
        
        # Get angle statistics for this window
        _, percentiles_skeleton = analyze_angle_changes(angle_changes_list)
        list_all = np.array(min_max_angles)  # Reference values from training
        list_one = np.array(percentiles_skeleton)  # Values for this window

        # Calculate scaling factors
        min_factor = np.minimum(list_all[:, 2] / list_one[:, 2], 1)  # P5 factors, capped at 1
        max_factor = np.maximum(list_all[:, 6] / list_one[:, 6], 1)  # P95 factors, floored at 1

        # Adjust head segment (indices 0-4) to use same scaling factor
        max_val = np.max(min_factor[0:5])
        min_factor[0:5] = max_val
        min_val = np.min(max_factor[0:5])
        max_factor[0:5] = min_val
        
        # Store basic information
        segment_key = 'top-k' if is_top_k else 'non-top-k'
        perturb_results.append({
            'filename': filename, 
            'window': window_number,
            segment_key: joints_to_analyze,
            'min_angles': min_factor,
            'max_angles': max_factor,
            'orig_pred': orig_pred,
        })

        # Reduce angles (min perturbation)
        step = 'min'
        for factor in factors_min:
            new_frames = copy.deepcopy(frames_in_window)
            angle_deltas = min_factor * factor
            
            # Perform angle perturbation
            new_frames = perturb_angle(
                new_frames,
                angle_deltas,
                j=joints_to_analyze,
                mode=step,
                debug=False
            )
                
            # Predict with perturbed data
            _, new_preds, _ = predict_with_pvb(
                tracking_coords=new_frames, 
                body_parts=BODY_PARTS, 
                frame_rate=frame_rate, 
                total_frames=new_frames.shape[0],
                pred_frame_rate=30.0,
                pred_interval_seconds=2.5,
                window_stride=2,
                num_models=10,
                num_portions=7,
                prediction_threshold=CP_RISK_THRESHOLD,
                xai_technique='cam'
            )
            new_pred = np.median(new_preds)
            update_list(perturb_results, filename, window_number, factor, new_pred, step)

        # Increase angles (max perturbation)
        step = 'max'
        for factor in factors_max:
            new_frames = copy.deepcopy(frames_in_window)
            angle_deltas = max_factor * factor
            
            # Perform angle perturbation
            new_frames = perturb_angle(
                new_frames,
                angle_deltas,
                j=joints_to_analyze,
                mode=step,
                debug=False
            )
                
            # Predict with perturbed data
            _, new_preds, _ = predict_with_pvb(
                tracking_coords=new_frames, 
                body_parts=BODY_PARTS, 
                frame_rate=frame_rate, 
                total_frames=new_frames.shape[0],
                pred_frame_rate=30.0,
                pred_interval_seconds=2.5,
                window_stride=2,
                num_models=10,
                num_portions=7,
                prediction_threshold=CP_RISK_THRESHOLD,
                xai_technique='cam'
            )
            new_pred = np.median(new_preds)
            update_list(perturb_results, filename, window_number, factor, new_pred, step)

    # Save results to pickle and Excel
    with open(f"{output_filename}.pkl", 'wb') as file:
        pickle.dump(perturb_results, file)
    
    save_to_excel_with_formatting(perturb_results, output_filename, CP_RISK_THRESHOLD)


def main():
    """
    Main execution function for angle perturbation analysis.
    """
    # Path configuration
    data_path = "./data/"  # Folder containing tracked CSV files for joint positions
    # json_file_path - Path to JSON file with video frame rates dictionary
    json_file_path = './data/video_fps_dict.json'
    # min_max_angles_path - Path to pickled file with reference angle statistics
    min_max_angles_path = './data/min_max_angles_trainval.pkl'
    # filename_window_path - Path to pickled file with window data
    filename_window_path = './data/filename_window_lowrisk.pkl'
    
    # Load data
    video_fps_dict = load_video_fps_dict(json_file_path)
    
    with open(min_max_angles_path, 'rb') as f:
        min_max_angles_trainval = pickle.load(f)
    
    with open(filename_window_path, 'rb') as f:
        filename_window_lowrisk = pickle.load(f)
    
    print(f"There are {len(filename_window_lowrisk)} windows to process")
    
    # Run perturbation analysis for non-top-k joints
    run_angle_perturbation(
        filename_windows=filename_window_lowrisk,
        joints_to_analyze=NON_TOP_K_LOW_RISK,
        is_top_k=False,
        data_path=data_path,
        output_filename="cam_nontopk_lowrisk_test",
        min_max_angles=min_max_angles_trainval,
        video_fps_dict=video_fps_dict
    )
    
    # Run perturbation analysis for top-k joints
    run_angle_perturbation(
        filename_windows=filename_window_lowrisk,
        joints_to_analyze=TOP_K_LOW_RISK,
        is_top_k=True,
        data_path=data_path,
        output_filename="cam_topk_lowrisk_test",
        min_max_angles=min_max_angles_trainval,
        video_fps_dict=video_fps_dict
    )


if __name__ == "__main__":
    main()