"""
XAI-informed Perturbation Analysis Module.

This module implements perturbation-based analysis for understanding
the behavior of our Cerebral Palsy prediction model.
"""

import os
import math
import json
import pickle
import copy
import random
from collections import defaultdict
from typing import List, Dict, Tuple, Any, Optional, Union

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from scipy.interpolate import interp1d
from openpyxl import load_workbook
from openpyxl.styles import Font

# Import prediction function from private module
from predict import predict, predict_with_pvb
from utils.predict_helpers import read_csv_to_array


# Constants
BODY_PARTS = [
    'head_top', 'nose', 'right_ear', 'left_ear', 'upper_neck',
    'right_shoulder', 'right_elbow', 'right_wrist', 'thorax',
    'left_shoulder', 'left_elbow', 'left_wrist', 'pelvis',
    'right_hip', 'right_knee', 'right_ankle', 'left_hip',
    'left_knee', 'left_ankle'
]

CP_RISK_THRESHOLD = 0.35025954  # Risk threshold for CP classification
CAM_THRESHOLD = 0.2915886665789474  # CAM threshold

# Joint groupings
TOP_K_LOW_RISK = [6, 7, 10, 11, 14, 15]  # Identified important joints
NON_TOP_K_LOW_RISK = [0, 1, 2, 3, 4, 5, 8, 9, 12, 13, 16, 17, 18]  # Other joints

# Body segment definitions for coherent perturbation
TOP_K_SEGMENTS = [[6, 7], [10, 11], [14, 15]]
NON_TOP_K_SEGMENTS = [[0, 1, 2, 3, 4], [5, 8, 9], [12, 13, 16], [17, 18]]


def load_video_fps_dict(json_path: str) -> Dict[str, float]:
    """
    Load video frame rate dictionary from JSON file and modify keys.
    
    Args:
        json_path: Path to the JSON file containing video FPS information
        
    Returns:
        Dictionary mapping CSV filenames to video frame rates
    """
    with open(json_path, 'r') as json_file:
        video_fps_dict = json.load(json_file)
    
    # Add prefix 'tracked_' and suffix '.csv' to all keys
    return {'tracked_' + key + '.csv': value for key, value in video_fps_dict.items()}


def classify_data_by_filename_and_windows(
    data: Dict[str, Dict[int, Dict[str, Any]]],
    threshold: float
) -> Dict[int, List[Dict[str, Any]]]:
    """
    Classify data into low-risk (0) and high-risk (1) groups based on CP risk.
    
    Args:
        data: Nested dictionary with filenames and window data
        threshold: Risk threshold for classification
        
    Returns:
        Dictionary with classified data grouped by risk level (0 or 1)
    """
    grouped_data = {0: [], 1: []}  # Group data only by 0 and 1

    # Step 1: Classify filenames based on the median of all window_cp_risks
    for filename, windows_data in data.items():
        all_window_cp_risks = []

        # Collect the median cp_risks for all windows in the file
        for window_num, window_data in windows_data.items():
            y = window_data['y']
            window_cp_risks = np.median(np.asarray(y), axis=0)
            all_window_cp_risks.append(window_cp_risks)

        # Calculate overall percentiles for the filename
        filename_median = np.median(all_window_cp_risks)
        filename_lower, filename_upper = np.percentile(all_window_cp_risks, [25, 75])

        # Classify the filename into group 0 or 1 based on percentiles
        if filename_upper < threshold:
            # Filename belongs to group 0 (low risk)
            filename_group = 0
        elif filename_lower >= threshold:
            # Filename belongs to group 1 (high risk)
            filename_group = 1
        else:
            # Filename has mixed risk levels; skip it
            continue

        # Step 2: Classify individual windows based on risk level
        for window_num, window_data in windows_data.items():
            y = window_data['y']
            
            # Compute window_cp_risks for classification
            window_cp_risks = np.median(np.asarray(y), axis=0)
            window_cp_risks_lower, window_cp_risks_upper = np.percentile(np.asarray(y), [25, 75])
            
            if filename_group == 0:
                # Logic for group 0 (low risk)
                if window_cp_risks_upper < threshold:
                    # Retain the window in group 0
                    grouped_data[0].append({
                        'x': window_data['x'],             # Kinematic data
                        'y': window_data['y'],             # Prediction data
                        'cams': window_data['cams'],       # CAMs data
                        'filename': filename,              # The source filename
                        'window_number': window_num        # The window number
                    })
            elif filename_group == 1:
                # Logic for group 1 (high risk)
                if window_cp_risks_lower >= threshold:
                    # Retain the window in group 1
                    grouped_data[1].append({
                        'x': window_data['x'],             # Kinematic data
                        'y': window_data['y'],             # Prediction data
                        'cams': window_data['cams'],       # CAMs data
                        'filename': filename,              # The source filename
                        'window_number': window_num        # The window number
                    })
    
    print('grouped_data[0] : ', len(grouped_data[0]))
    print('grouped_data[1] : ', len(grouped_data[1]))

    return grouped_data


def select_representative_windows(
    grouped_data: Dict[int, List[Dict[str, Any]]],
    num_samples: int
) -> List[int]:
    """
    Select representative windows while maintaining proportional sampling by file.
    
    Args:
        grouped_data: Dictionary with classified data
        num_samples: Number of windows to select
        
    Returns:
        List of selected window indices
    """
    # First, organize data by filename to count windows per file
    filename_windows = defaultdict(list)
    for idx, data in enumerate(grouped_data[0]):
        filename_windows[data['filename']].append(idx)

    # Calculate the total number of windows and windows per filename
    total_windows = len(grouped_data[0])
    windows_per_file = {fname: len(indices) for fname, indices in filename_windows.items()}

    # Calculate the number of samples to take per filename
    file_proportions = {
        fname: int((count / total_windows) * num_samples)
        for fname, count in windows_per_file.items()
    }

    # Ensure we're not losing samples due to rounding
    remaining_samples = num_samples - sum(file_proportions.values())
    if remaining_samples > 0:
        # Add remaining samples to files with highest window counts
        sorted_files = sorted(windows_per_file.items(), key=lambda x: x[1], reverse=True)
        for i in range(remaining_samples):
            file_proportions[sorted_files[i][0]] += 1

    # Randomly select windows while maintaining proportions
    selected_windows = []
    for filename, num_to_select in file_proportions.items():
        available_indices = filename_windows[filename]
        selected = random.sample(available_indices, min(num_to_select, len(available_indices)))
        selected_windows.extend(selected)

    return selected_windows


def calculate_min_scaling(
    vx: np.ndarray,
    vy: np.ndarray,
    joints: List[int],
    perc_pos: Dict[str, List[float]]
) -> List[List]:
    """
    Calculate minimum scaling factors to align velocities with dataset percentiles.
    
    Args:
        vx: X-axis velocity data
        vy: Y-axis velocity data
        joints: List of joint indices to analyze
        perc_pos: Dictionary with positive velocity percentiles
        
    Returns:
        List of scaling information for each joint
    """
    results = []
    for joint in joints:
        # Filter non-zero velocities
        vx_non_zero = vx[vx[:, joint] != 0, joint]
        if vx_non_zero.size == 0:
            vx_non_zero = vx
        vy_non_zero = vy[vy[:, joint] != 0, joint]
        if vy_non_zero.size == 0:
            vy_non_zero = vy

        # Calculate 5th percentile of sample velocities
        vx_pos_5th_sample = np.percentile(np.abs(vx_non_zero), 5)
        vy_pos_5th_sample = np.percentile(np.abs(vy_non_zero), 5)

        # Get the 5th percentile for the overall dataset
        vx_pos_5th_data = perc_pos[f'vx{joint}'][0]
        vy_pos_5th_data = perc_pos[f'vy{joint}'][0]

        # Calculate min scaling factor
        vx_min_scaling = vx_pos_5th_data / vx_pos_5th_sample
        vy_min_scaling = vy_pos_5th_data / vy_pos_5th_sample

        # Append results
        results.append([f'vx{joint}', round(vx_min_scaling, 4), round(vx_pos_5th_sample, 6), round(vx_pos_5th_data, 6)])
        results.append([f'vy{joint}', round(vy_min_scaling, 4), round(vy_pos_5th_sample, 6), round(vy_pos_5th_data, 6)])

    return results


def calculate_max_scaling(
    vx: np.ndarray,
    vy: np.ndarray,
    joints: List[int],
    perc_all: Dict[str, List[float]]
) -> List[List]:
    """
    Calculate maximum scaling factors to align velocities with dataset percentiles.
    
    Args:
        vx: X-axis velocity data
        vy: Y-axis velocity data
        joints: List of joint indices to analyze
        perc_all: Dictionary with velocity percentiles
        
    Returns:
        List of scaling information for each joint
    """
    results = []
    for joint in joints:
        # Filter non-zero velocities
        vx_non_zero = vx[vx[:, joint] != 0, joint]
        if vx_non_zero.size == 0:
            vx_non_zero = vx
        vy_non_zero = vy[vy[:, joint] != 0, joint]
        if vy_non_zero.size == 0:
            vy_non_zero = vy

        # Calculate the 5th and 95th percentiles for sample velocities
        vx_sample_5th, vx_sample_95th = np.percentile(vx_non_zero, [5, 95])
        vy_sample_5th, vy_sample_95th = np.percentile(vy_non_zero, [5, 95])

        # Determine the sample velocity limit for max scaling
        vx_sample_limit = max(abs(vx_sample_5th), abs(vx_sample_95th))
        vy_sample_limit = max(abs(vy_sample_5th), abs(vy_sample_95th))

        # Get the 5th and 95th percentiles for the overall dataset
        vx_data_limit = max(abs(perc_all[f'vx{joint}'][0]), abs(perc_all[f'vx{joint}'][1]))
        vy_data_limit = max(abs(perc_all[f'vy{joint}'][0]), abs(perc_all[f'vy{joint}'][1]))

        # Calculate max scaling factor
        vx_max_scaling = vx_data_limit / vx_sample_limit
        vy_max_scaling = vy_data_limit / vy_sample_limit

        # Append results
        results.append([f'vx{joint}', round(vx_max_scaling, 4), round(vx_sample_limit, 6), round(vx_data_limit, 6)])
        results.append([f'vy{joint}', round(vy_max_scaling, 4), round(vy_sample_limit, 6), round(vy_data_limit, 6)])

    return results


def analyze_window_with_adjusted_scaling(
    data: Dict[int, List[Dict[str, Any]]],
    window_num: int,
    label: int,
    top_k: List[int],
    non_top_k: List[int],
    perc_all: Dict[str, List[float]],
    perc_pos: Dict[str, List[float]]
) -> Tuple[List[List], List[List], str, int]:
    """
    Analyze a window and calculate adjusted scaling factors for velocity perturbation.
    
    Args:
        data: Dictionary with classified data
        window_num: Window number to analyze
        label: Risk label (0 or 1)
        top_k: List of important joint indices
        non_top_k: List of other joint indices
        perc_all: Dictionary with overall velocity percentiles
        perc_pos: Dictionary with positive velocity percentiles
        
    Returns:
        Tuple with scaling results for top_k and non_top_k joints, filename, and window number
    """
    all_joints = top_k + non_top_k  # Ensures all joints are evaluated

    if label not in data or len(data[label]) == 0:
        print(f"No data found for label {label}")
        return [], [], "", 0

    # Select window from specified group
    selected_window = data[0][window_num]
    
    # Extract velocity components and metadata
    vx = selected_window['x'][2]  # vx corresponds to index 2 in 'x' (shape: 150, 19)
    vy = selected_window['x'][3]  # vy corresponds to index 3 in 'x' (shape: 150, 19)
    filename = selected_window['filename']
    window_number = selected_window['window_number']
    
    print(f"Filename: {filename}")
    print(f"Window Number: {window_number}")
    
    # Calculate min and max scaling factors
    min_scaling_results = calculate_min_scaling(vx, vy, all_joints, perc_pos)
    max_scaling_results = calculate_max_scaling(vx, vy, all_joints, perc_all)

    # Combine min and max results for each group
    top_k_results = []
    non_top_k_results = []
    
    for i, joint in enumerate(all_joints):
        # For each joint, handle both vx and vy components separately
        for component in ['vx', 'vy']:
            index = 2 * i if component == 'vx' else 2 * i + 1  # 2*i for vx, 2*i+1 for vy
            joint_name = f"{component}{joint}"
            min_scale = min_scaling_results[index][1]
            max_scale = max_scaling_results[index][1]
            
            # Ensure min_scale doesn't exceed 1 and max_scale is not less than 1
            min_scale = min(1, min_scale)
            max_scale = max(1, max_scale)
            
            result_row = [
                joint_name, min_scale, min_scaling_results[index][2], min_scaling_results[index][3],
                max_scale, max_scaling_results[index][2], max_scaling_results[index][3]
            ]
            
            # Determine which list to append to based on joint membership
            if joint in top_k:
                top_k_results.append(result_row)
            elif joint in non_top_k:
                non_top_k_results.append(result_row)

    # Create a dictionary to store vx and vy for each joint
    joint_dict = {}
    for results in [top_k_results, non_top_k_results]:
        for row in results:
            joint_name = row[0]
            joint_number = int(joint_name[2:])  # Extract the joint number
            if joint_number not in joint_dict:
                joint_dict[joint_number] = {}
            joint_dict[joint_number][joint_name[:2]] = row  # 'vx' or 'vy' as key
        
    # Adjust min and max scaling for each joint by comparing vx and vy
    for joint, components in joint_dict.items():
        if 'vx' in components and 'vy' in components:
            vx_row = components['vx']
            vy_row = components['vy']
            
            # Adjust min scaling by taking the maximum of vx and vy min scaling factors
            adjusted_min_scale = max(vx_row[1], vy_row[1])
            vx_row[1] = vy_row[1] = adjusted_min_scale
            
            # Adjust max scaling by taking the minimum of vx and vy max scaling factors
            adjusted_max_scale = min(vx_row[4], vy_row[4])
            vx_row[4] = vy_row[4] = adjusted_max_scale

    # Define segments for further adjustment across joints in the same segment
    segments = [
        [0, 1, 2, 3, 4], [5, 8, 9], [6, 7], [10, 11], [12, 13, 16], [14, 15], [17, 18]
    ]
    
    # Adjust scaling factors across segments
    for segment in segments:
        # Check if any joint in segment exists in joint_dict
        segment_joints = [j for j in segment if j in joint_dict]
        if not segment_joints:
            continue
            
        # Collect min and max scaling values across the entire segment
        segment_min_scale = max(joint_dict[joint]['vx'][1] for joint in segment_joints)
        segment_max_scale = min(joint_dict[joint]['vx'][4] for joint in segment_joints)
        
        # Apply the unified scaling across all joints in the segment
        for joint in segment_joints:
            joint_dict[joint]['vx'][1] = joint_dict[joint]['vy'][1] = segment_min_scale
            joint_dict[joint]['vx'][4] = joint_dict[joint]['vy'][4] = segment_max_scale
    
    # Prepare unified top_k and non_top_k result lists without vx/vy distinction
    unified_top_k_results = []
    unified_non_top_k_results = []

    for joint, components in joint_dict.items():
        if 'vx' not in components:
            continue
            
        vx_row = components['vx']
        
        # Unified row without vx/vy distinction (joint number only)
        unified_row = [
            joint, vx_row[1], vx_row[2], vx_row[3], 
            vx_row[4], vx_row[5], vx_row[6]
        ]
        
        # Append to the appropriate list based on whether joint is in top_k or non_top_k
        if joint in top_k:
            unified_top_k_results.append(unified_row)
        elif joint in non_top_k:
            unified_non_top_k_results.append(unified_row)

    return unified_top_k_results, unified_non_top_k_results, filename, window_number


def get_frame_range_for_window(
    window_number: int,
    frame_rate: float,
    num_frames: int,
    pred_interval_seconds: float = 5
) -> Tuple[Optional[int], Optional[int]]:
    """
    Calculate the start and end frame indices for a specific window.
    
    Args:
        window_number: Index of the window to analyze
        frame_rate: Frame rate of the video
        num_frames: Total number of frames in the video
        pred_interval_seconds: Size of prediction window in seconds
        
    Returns:
        Tuple with start and end frame indices, or (None, None) if invalid
    """
    window_size = int(frame_rate * pred_interval_seconds)
    step_size = int(window_size / 2)
    
    # Compute start and end frame
    start_frame = window_number * step_size
    end_frame = start_frame + window_size - 1
    
    # Ensure that the start frame doesn't go beyond the total number of frames
    if start_frame >= num_frames:
        return None, None  # Invalid window number
    
    # Ensure that the end frame does not exceed the total number of frames
    if end_frame >= num_frames:
        end_frame = num_frames - 1
    
    # Adjust the window if it's the last valid window
    # Check if this window is the last full window or partial window
    total_windows = math.floor((num_frames - step_size) / step_size) + 1
    if window_number == total_windows - 1:
        start_frame = num_frames - window_size if num_frames >= window_size else 0
        end_frame = num_frames - 1
    
    return start_frame, end_frame


def modify_speed(
    X: np.ndarray,
    s: float,
    body_part_indices: List[int]
) -> np.ndarray:
    """
    Modify the speed of motion for specified body parts.
    
    Args:
        X: Array of shape (frames, 38) containing xy coordinates
        s: Speed factor (>1 speeds up, <1 slows down)
        body_part_indices: List of body part indices to modify
        
    Returns:
        Modified array with adjusted motion speed
    """
    # X has shape (frames, 38)
    X = np.asarray(X, dtype=np.float64)
    frames, n_points = X.shape
    assert n_points == 38, "Expected 38 points (19 body parts with x, y coordinates)"
    
    # Create a copy of X to modify the data
    X_modified = X.copy()

    # Function to compress (speed up) or stretch (slow down) the motion
    def interpolate_and_modify_trajectory(original_trajectory, factor, total_frames):
        original_length = len(original_trajectory)
        new_length = int(original_length / factor)

        # Create the interpolation function
        f_interp = interp1d(np.linspace(0, 1, original_length), original_trajectory, kind='linear')

        # Create a compressed or stretched version of the trajectory
        new_trajectory = f_interp(np.linspace(0, 1, new_length))

        # If speeding up (s > 1), reflect the trajectory when running out of data
        if factor > 1:
            while len(new_trajectory) < total_frames:
                # Reflect the trajectory to ensure smooth motion when "running out" of trajectory
                new_trajectory = np.concatenate([new_trajectory, new_trajectory[::-1]])

            # Slice to match the number of frames
            return new_trajectory[:total_frames]

        # If slowing down (s < 1), stretch the trajectory
        elif factor < 1:
            if new_length >= total_frames:
                # If the new trajectory is longer than the total frames, truncate it
                return new_trajectory[:total_frames]
            else:
                # Stretch the trajectory and pad with the last value if needed
                return np.pad(new_trajectory, (0, total_frames - new_length), mode='edge')

        return original_trajectory  # If no change in speed (s=1), return original trajectory
    
    # Loop through the specified body parts to modify
    for part_idx in body_part_indices:
        # For each body part, modify both x and y coordinates
        x_idx = 2 * part_idx     # x coordinate index
        y_idx = 2 * part_idx + 1 # y coordinate index

        # Modify the x and y trajectories
        X_modified[:, x_idx] = interpolate_and_modify_trajectory(X[:, x_idx], s, frames)
        X_modified[:, y_idx] = interpolate_and_modify_trajectory(X[:, y_idx], s, frames)

    X_modified = X_modified.astype('<U19')

    return X_modified


def convert_to_nested(
    segments: List[List[int]],
    scaling_values: List[float]
) -> List[List[float]]:
    """
    Convert a flat list of scaling values to a nested list matching segment structure.
    
    Args:
        segments: List of joint segment groups
        scaling_values: Flat list of scaling values
        
    Returns:
        Nested list of scaling values grouped by segment
    """
    nested_scale = []
    index = 0
    for segment in segments:
        segment_length = len(segment)
        nested_scale.append(scaling_values[index:index + segment_length])
        index += segment_length
    return nested_scale


def update_list(
    perturb_list: List[Dict[str, Any]],
    filename: str,
    window_number: int,
    factor: float,
    new_pred: float,
    step: str
) -> None:
    """
    Update perturbation results in the results list.
    
    Args:
        perturb_list: List of perturbation results
        filename: Video filename
        window_number: Window index
        factor: Speed factor applied
        new_pred: New prediction value
        step: Direction of perturbation ('min' or 'max')
    """
    # Check if the entry already exists
    for entry in perturb_list:
        if entry['filename'] == filename and entry['window'] == window_number:
            # If found, update the existing entry
            key = '*' + step
            factor_str = f'{factor:.2f}' if factor < 1 else f'{int(factor)}'
            entry[f'{factor_str}{key}'] = new_pred
            return


def save_to_excel_with_formatting(
    data: List[Dict[str, Any]],
    filename: str,
    threshold: float
) -> None:
    """
    Save data to Excel with conditional formatting for risk values.
    
    Args:
        data: List of dictionaries with perturbation results
        filename: Output filename without extension
        threshold: Risk threshold for highlighting
    """
    # Convert list of dictionaries to DataFrame
    df = pd.DataFrame(data)
    
    # Save DataFrame to Excel
    excel_file = f"{filename}.xlsx"
    df.to_excel(excel_file, index=False)
    
    # Load the workbook and select the active sheet
    wb = load_workbook(excel_file)
    ws = wb.active

    # Apply conditional formatting: if value > threshold, turn the text red
    red_font = Font(color="FF0000")  # Red text
    for col in ws.iter_cols(min_row=2, min_col=2, max_col=ws.max_column):
        for cell in col:
            if isinstance(cell.value, (int, float)) and cell.value > threshold:
                cell.font = red_font

    # Save the workbook with formatting
    wb.save(excel_file)


def run_perturbation_analysis(
    grouped_data: Dict[int, List[Dict[str, Any]]],
    selected_windows: List[int],
    segments: List[List[int]],
    is_top_k: bool,
    data_path: str,
    output_filename: str,
    video_fps_dict: Dict[str, float]
) -> None:
    """
    Run perturbation analysis on selected windows for specified joint segments.
    
    Args:
        grouped_data: Dictionary with classified data
        selected_windows: List of window indices to analyze
        segments: List of joint segment groups to perturb
        is_top_k: Whether analyzing top-k (True) or non-top-k (False) joints
        data_path: Path to tracking data files
        output_filename: Base filename for output files
        video_fps_dict: Dictionary with video frame rates
    """
    # Factors for min and max perturbation
    factors_min = [0.2, 0.25, 1/3, 0.5, 1]
    factors_max = [1, 2, 3, 4, 5]
    
    perturb_results = []
    pred_seconds = 5
    
    for window_idx in selected_windows:
        window_data = data_windows[window_idx]
        filename = window_data['filename']
        window_number = int(window_data['window'])
        
        # Load tracking data for this file
        csv_path = os.path.join(data_path, filename)
        tracking_coords = np.array(read_csv_to_array(csv_path))
        num_frames = tracking_coords.shape[0]
        frame_rate = video_fps_dict.get(filename)
        
        if frame_rate is None:
            print(f"Warning: No frame rate found for {filename}, skipping.")
            continue
        
        # Get frame range for this window
        start_frame, end_frame = get_frame_range_for_window(
            window_number, frame_rate, num_frames, pred_seconds
        )
        
        if start_frame is None:
            print(f"Invalid window {window_number} for {filename}, skipping.")
            continue
            
        print(f"Window {window_number} from {filename}: frames {start_frame} to {end_frame}")
        
        # Extract frames for this window
        frames_in_window = tracking_coords[start_frame:end_frame + 1]

        # Get original prediction
        _, orig_preds, _ = predict_with_pvb(
            tracking_coords=frames_in_window, 
            body_parts=BODY_PARTS, 
            frame_rate=frame_rate, 
            total_frames=frames_in_window.shape[0],
            pred_frame_rate=30.0,
            pred_interval_seconds=2.5,
            window_stride=2,
            num_models=10,
            num_portions=7,
            prediction_threshold=CP_RISK_THRESHOLD,
            xai_technique='cam'
        )
        orig_pred = np.median(orig_preds)
        
        # Get scaling factors
        if is_top_k:
            min_scale = window_data['min_topk']
            max_scale = window_data['max_topk']
            nested_min = convert_to_nested(segments, min_scale)
            nested_max = convert_to_nested(segments, max_scale)
        else:
            min_scale = window_data['min_non_topk']
            max_scale = window_data['max_non_topk']
            nested_min = convert_to_nested(segments, min_scale)
            nested_max = convert_to_nested(segments, max_scale)
        
        # Store basic information
        segment_key = 'top-k' if is_top_k else 'non-top-k'
        perturb_results.append({
            'filename': filename,
            'window': window_number,
            segment_key: segments,
            'min_speed': nested_min,
            'max_speed': nested_max,
            'orig_pred': orig_pred,
        })
        
        # Slow down (min) perturbation
        step = 'min'
        for factor in factors_min:
            new_frames = copy.deepcopy(frames_in_window)
            
            for segment, scale_values in zip(segments, nested_min):
                if not scale_values:
                    continue
                    
                joints_to_perturb = segment
                speed = scale_values[0] * factor
                new_frames = modify_speed(new_frames, speed, joints_to_perturb)
                
            # Predict with perturbed data
            _, new_preds, _ = predict_with_pvb(
                tracking_coords=new_frames, 
                body_parts=BODY_PARTS, 
                frame_rate=frame_rate, 
                total_frames=new_frames.shape[0],
                pred_frame_rate=30.0,
                pred_interval_seconds=2.5,
                window_stride=2,
                num_models=10,
                num_portions=7,
                prediction_threshold=CP_RISK_THRESHOLD,
                xai_technique='cam'
            )
            new_pred = np.median(new_preds)
            update_list(perturb_results, filename, window_number, factor, new_pred, step)
        
        # Speed up (max) perturbation
        step = 'max'
        for factor in factors_max:
            new_frames = copy.deepcopy(frames_in_window)
            
            for segment, scale_values in zip(segments, nested_max):
                if not scale_values:
                    continue
                    
                joints_to_perturb = segment
                speed = scale_values[0] * factor
                new_frames = modify_speed(new_frames, speed, joints_to_perturb)
                
            # Predict with perturbed data
            _, new_preds, _ = predict_with_pvb(
                tracking_coords=new_frames, 
                body_parts=BODY_PARTS, 
                frame_rate=frame_rate, 
                total_frames=new_frames.shape[0],
                pred_frame_rate=30.0,
                pred_interval_seconds=2.5,
                window_stride=2,
                num_models=10,
                num_portions=7,
                prediction_threshold=CP_RISK_THRESHOLD,
                xai_technique='cam'
            )
            new_pred = np.median(new_preds)
            update_list(perturb_results, filename, window_number, factor, new_pred, step)
    
    # Save results to pickle and Excel
    with open(f"{output_filename}.pkl", 'wb') as file:
        pickle.dump(perturb_results, file)
    
    save_to_excel_with_formatting(perturb_results, output_filename, CP_RISK_THRESHOLD)


def main():
    """
    Main execution function for perturbation analysis.
    """
    # Path configuration
    data_path = "./data/"  # Folder containing tracked CSV files for joint positions
    # json_file_path - Path to JSON file with video frame rates dictionary
    json_file_path = './data/video_fps_dict.json'
    # test_data_path - Path to pickled file containing test data with CAM values and PVB input data
    test_data_path = './data/cams_pvb_test.pkl'
    # percentiles_path - Path to pickled file with velocity percentiles data
    percentiles_path = './data/percentiles_cam_lowrisk_trainval.pkl'
    
    # Load data
    video_fps_dict = load_video_fps_dict(json_file_path)
    
    with open(test_data_path, 'rb') as f:
        cams_pvb_test = pickle.load(f)
    
    print('Number of test files:', len(cams_pvb_test.keys()))
    
    # Create a deep copy to avoid modifying original data
    data_copy = copy.deepcopy(cams_pvb_test)
    grouped_data = classify_data_by_filename_and_windows(data_copy, CP_RISK_THRESHOLD)
    
    # Load or calculate percentiles
    with open(percentiles_path, 'rb') as f:
        vpercentile_all, vpercentile_pos = pickle.load(f)
    
    # Select representative windows or load previously selected windows
    try:
        with open('selected_windows_lowrisk.pkl', 'rb') as f:
            selected_windows = pickle.load(f)
    except FileNotFoundError:
        selected_windows = select_representative_windows(grouped_data, 493)
        with open('selected_windows_lowrisk.pkl', 'wb') as f:
            pickle.dump(selected_windows, f)
    
    # Calculate window data for selected windows
    global data_windows
    data_windows = []
    
    for n in selected_windows:
        top_k_results, non_top_k_results, file, window = analyze_window_with_adjusted_scaling(
            grouped_data, 
            window_num=n,
            label=0, 
            top_k=TOP_K_LOW_RISK, 
            non_top_k=NON_TOP_K_LOW_RISK, 
            perc_all=vpercentile_all, 
            perc_pos=vpercentile_pos
        )
        
        data_windows.append({
            'filename': file, 
            'window': window, 
            'min_topk': [row[1] for row in top_k_results],
            'max_topk': [row[4] for row in top_k_results],
            'min_non_topk': [row[1] for row in non_top_k_results],
            'max_non_topk': [row[4] for row in non_top_k_results],
        })
    
    # Run perturbation analysis for top-k segments
    run_perturbation_analysis(
        grouped_data=grouped_data,
        selected_windows=selected_windows,
        segments=TOP_K_SEGMENTS,
        is_top_k=True,
        data_path=data_path,
        output_filename="cam_topk_lowrisk_test",
        video_fps_dict=video_fps_dict
    )
    
    # Run perturbation analysis for non-top-k segments
    run_perturbation_analysis(
        grouped_data=grouped_data,
        selected_windows=selected_windows,
        segments=NON_TOP_K_SEGMENTS,
        is_top_k=False,
        data_path=data_path,
        output_filename="cam_nontopk_lowrisk_test",
        video_fps_dict=video_fps_dict
    )


if __name__ == "__main__":
    main()